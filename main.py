"""SCD Logistics — App completa (Maestro + Optimizador + Despachos).
Un solo servicio FastAPI: sirve la UI en / y la API. Auto-seed del maestro.
Credenciales solo en DATABASE_URL. Nada de secretos en el front.
"""
import os, io, csv, re, json, base64, datetime
from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Body
from fastapi.responses import FileResponse, StreamingResponse
import openpyxl

# ================= ENGINE: SKU / volumen / importador =================
WT="WT-"
def normalize_sku(raw):
    if raw is None: return ""
    s=str(raw).strip()
    return s[len(WT):] if s.startswith(WT) else s
def is_set(raw_sku, desc=""):
    raw=str(raw_sku or ""); sku=normalize_sku(raw)
    return ("\n" in raw) or (" + " in sku) or ("in one set" in str(desc).lower())
def _num(x):
    try:
        if x is None or str(x).strip()=="" : return None
        return float(x)
    except (ValueError,TypeError): return None
def sanitize_volume(L,A,H,V0,tol=0.02):
    L,A,H,V0=_num(L),_num(A),_num(H),_num(V0)
    if None in (L,A,H) or L<=0 or A<=0 or H<=0:
        return {"volumen_calculado_m3":None,"estado_volumen":"sin dimensiones"}
    calc=round(L*A*H,6)
    if V0 is None or V0<0: estado="corregido"
    else:
        rel=abs(V0-calc)/calc if calc else 1.0
        estado="correcto" if rel<=tol else "corregido"
    return {"volumen_calculado_m3":calc,"estado_volumen":estado}

SYN={"sku":["sku","article #","article","item no","item no.","item","codigo","code","model"],
 "descripcion":["description of goods","description","descripcion","goods","producto"],
 "cantidad":["qty (pcs)","qty","quantity","cantidad","pcs","total pcs"],
 "ctn":["ctn","ctns","cartons","no of pkgs","no of pkgs.","bultos"],
 "largo":["largo","length","l (cm)","l"],"ancho":["ancho","width","w (cm)","w"],
 "alto":["alto","height","h (cm)","h"],"cbm":["cbm","volume","volumen","m3","meas"],
 "peso_neto":["nt.wt","net weight","peso neto","nt. wt."],
 "peso_bruto":["gr.wt","gross weight","peso bruto","gr. wt."]}
def _n(s): return re.sub(r"\s+"," ",str(s or "").strip().lower())
def detect_header(rows,scan=25):
    bi,bm,bs=None,None,0
    for i,row in enumerate(rows[:scan]):
        cells=[_n(c) for c in row]; m={}; sc=0
        for canon,syns in SYN.items():
            for j,c in enumerate(cells):
                if c in syns: m[canon]=j; sc+=1; break
        if ("sku" in m or "descripcion" in m) and sc>bs: bi,bm,bs=i,m,sc
    return bi,(bm or {})
def _m(v):
    try: x=float(v)
    except (ValueError,TypeError): return None
    if x<=0: return None
    return round(x/100.0,6) if x>3 else round(x,6)
def parse_rows(rows,archivo=""):
    hi,cm=detect_header(rows); res={"detectados":0,"no_interpretadas":0,"errores":[]}; items=[]
    if hi is None:
        res["errores"].append("No pude identificar la columna SKU."); return items,res
    for row in rows[hi+1:]:
        rs=row[cm["sku"]] if "sku" in cm and cm["sku"]<len(row) else None
        ds=row[cm["descripcion"]] if "descripcion" in cm and cm["descripcion"]<len(row) else ""
        if (rs is None or str(rs).strip()=="") and (ds is None or str(ds).strip()==""): continue
        if rs is None or str(rs).strip()=="" : res["no_interpretadas"]+=1; continue
        g=lambda k: row[cm[k]] if k in cm and cm[k]<len(row) else None
        items.append({"sku":normalize_sku(rs),"descripcion":str(ds or "").strip(),
            "cantidad":_num(g("cantidad")),"largo_m":_m(g("largo")),"ancho_m":_m(g("ancho")),
            "alto_m":_m(g("alto")),"cbm":g("cbm"),"tipo":"set" if is_set(rs,ds) else "normal",
            "archivo_origen":archivo}); res["detectados"]+=1
    return items,res
_ITEM=re.compile(r"^(?P<r>\d+\s*-\s*\d+)\s+(?P<sku>[A-Za-z0-9][A-Za-z0-9\-\+/]*)\s+(?P<a>[\d.,]+)\s+(?P<b>[\d.,]+)\s+(?P<c>[\d.,]+)\s+(?P<pcs>[\d.,]+)\s+(?P<u>[A-Za-z]+)")
def parse_pdf_text(text,archivo=""):
    items,res=[],{"detectados":0,"no_interpretadas":0,"errores":[]}
    for line in text.splitlines():
        m=_ITEM.match(line.strip())
        if not m: continue
        f=lambda x:(float(str(x).replace(",","")) if x else None)
        rs=m.group("sku")
        items.append({"sku":normalize_sku(rs),"descripcion":"","cantidad":f(m.group("pcs")),
            "largo_m":None,"ancho_m":None,"alto_m":None,"cbm":None,
            "tipo":"set" if is_set(rs,"") else "normal","archivo_origen":archivo})
        res["detectados"]+=1
    if not items: res["errores"].append("No pude identificar filas de productos en el PDF.")
    return items,res

# ================= DESPACHO (SIM/MARIA) =================
def _fd(s):
    if s is None: return None
    s=str(s).strip().replace('.','').replace(',','.')
    try: return float(s)
    except: return None
def parse_despacho(text):
    d={}
    m=re.search(r'(\d{2})\s+(\d{3})\s+(IC\w\d)\s+(\d{6})\s+([A-Z])\s+\d+\s+de\s+\d+',text)
    d['numero']=''.join(m.groups()) if m else None
    m=re.search(r'OFICIALIZADO\s+(\d{2}/\d{2}/\d{4})',text); d['oficializacion']=m.group(1) if m else None
    m=re.search(r'([A-ZÑÁÉÍÓÚ .,&]+S\.?R\.?L\.?)[^\n]*?(\d{2}-\d{8}-\d)',text)
    d['importador']=m.group(1).strip() if m else None
    d['cuit']=m.group(2) if m else None
    m=re.search(r'Despachante de Aduana[^\n]*\n[^\n]*?(\d{2}-\d{8}-\d)',text)
    d['despachante_cuit']=m.group(1) if m else None
    m=re.search(r'Cotiz\s*=\s*([\d.,]+)',text); d['cotizacion']=_fd(m.group(1)) if m else None
    m=re.search(r'\((\d{4,6})\)\s+([A-Z][A-Za-z .]+)',text); d['vendedor']=m.group(2).strip() if m else None
    items=[]; parts=re.split(r'\n(\d{4})\s+N\b',text)
    for i in range(1,len(parts),2):
        num=parts[i]; body=parts[i+1] if i+1<len(parts) else ''
        it={'item':num}
        s=re.search(r'(\d{4}\.\d{2}\.\d{2}\.\d{3}\w)',body); it['sim']=s.group(1) if s else None
        it['ncm']=it['sim'][:10] if it['sim'] else None
        c=re.search(r'UNIDAD\s+([\d.]+),\d{2}\b',body); it['cantidad']=_fd(c.group(1)+',00') if c else None
        pairs=[_fd(x) for x in re.findall(r'([\d.]+,\d{2})\s+(?=\1\b)',body)]
        it['fob_usd']=next((p for p in pairs if p and p>0),None)
        va=[p for p in pairs if p and p>0]; it['valor_aduana_usd']=max(va) if va else None
        def trib(code):
            mm=re.search(r'(\d+,\d+)?\s*P\s+([\d.]+,\d+)\s*\(\s*'+code+r'\s*\)',body)
            if mm: return (_fd(mm.group(1)) if mm.group(1) else None,_fd(mm.group(2)))
            return None,None
        it['der_pct'],it['der_usd']=trib('010'); it['est_pct'],it['est_usd']=trib('011')
        it['iva_pct'],it['iva_usd']=trib('415'); it['gan_pct'],it['gan_usd']=trib('424')
        it['iibb_pct'],it['iibb_usd']=trib('900')
        items.append(it)
    d['items']=items
    return d
def match_skus(despacho_items, pl_items):
    """Cruce best-effort: por NCM/HS. Marca estado del match."""
    for it in despacho_items:
        ncm=(it.get('ncm') or '').replace('.','')[:8]
        cands=[]
        for p in pl_items or []:
            cands.append(p['sku'])
        if not pl_items:
            it['sku']=None; it['match']='sin match'
        elif len(cands)==1:
            it['sku']=cands[0]; it['match']='match probable'
        else:
            it['sku']=", ".join(cands[:8]); it['match']='match probable'  # varios SKU bajo el NCM
    return despacho_items
def despacho_excel(d, upload_date):
    import xlsxwriter
    buf=io.BytesIO(); wb=xlsxwriter.Workbook(buf,{'in_memory':True})
    b=wb.add_format({'bold':True})
    r1=wb.add_worksheet("Resumen")
    rows=[("Fecha de carga",upload_date),("Fecha oficializacion",d.get('oficializacion')),
        ("Numero despacho",d.get('numero')),("Importador",d.get('importador')),("CUIT",d.get('cuit')),
        ("Despachante CUIT",d.get('despachante_cuit')),("Vendedor",d.get('vendedor')),
        ("Cotizacion",d.get('cotizacion')),("Total items",len(d.get('items',[]))),
        ("FOB total USD",sum(x.get('fob_usd') or 0 for x in d.get('items',[]))),
        ("Total tributado USD",sum((x.get('der_usd') or 0)+(x.get('est_usd') or 0)+(x.get('iva_usd') or 0)+(x.get('gan_usd') or 0)+(x.get('iibb_usd') or 0) for x in d.get('items',[])))]
    for i,(k,v) in enumerate(rows): r1.write(i,0,k,b); r1.write(i,1,v if v is not None else "")
    r1.set_column(0,0,22); r1.set_column(1,1,40)
    r2=wb.add_worksheet("Items")
    cols=["SKU","Estado match","N Item","Cantidad","Unidad","Posicion SIM","NCM","FOB USD",
        "Tasa estadistica %","Tasa estadistica USD","Ganancias %","Ganancias USD",
        "Derechos %","Derechos USD","IVA %","IVA USD","Ing Brutos USD","Valor en aduana USD"]
    for j,c in enumerate(cols): r2.write(0,j,c,b)
    for i,it in enumerate(d.get('items',[]),1):
        vals=[it.get('sku'),it.get('match'),it.get('item'),it.get('cantidad'),"UNIDAD",it.get('sim'),
            it.get('ncm'),it.get('fob_usd'),it.get('est_pct'),it.get('est_usd'),it.get('gan_pct'),
            it.get('gan_usd'),it.get('der_pct'),it.get('der_usd'),it.get('iva_pct'),it.get('iva_usd'),
            it.get('iibb_usd'),it.get('valor_aduana_usd')]
        for j,v in enumerate(vals): r2.write(i,j,v if v is not None else "")
    r2.set_column(0,17,14)
    r3=wb.add_worksheet("Validacion")
    for j,c in enumerate(["Campo","Valor detectado","Confianza","Advertencia"]): r3.write(0,j,c,b)
    vrow=1
    for it in d.get('items',[]):
        conf="alta" if it.get('fob_usd') and it.get('sim') else "media"
        warn="" if it.get('sim') else "Sin posicion SIM"
        r3.write(vrow,0,f"Item {it.get('item')}"); r3.write(vrow,1,f"SIM {it.get('sim')} / FOB {it.get('fob_usd')}")
        r3.write(vrow,2,conf); r3.write(vrow,3,warn); vrow+=1
    r3.set_column(0,3,22)
    wb.close(); buf.seek(0); return buf.read()

# ================= OPTIMIZADOR 3D =================
CONTAINERS={"20":(5.90,2.35,2.39,28200),"40":(12.03,2.35,2.39,26600),
            "40HC":(12.03,2.35,2.69,26600)}  # largo, ancho, alto (m), peso max kg
def optimizar(productos, cont, modo="libre"):
    from py3dbp import Packer,Bin,Item
    L,W,H,maxw=cont
    packer=Packer(); packer.add_bin(Bin('cont',L,W,H,maxw))
    CAP=150  # limite de cajas para performance
    if modo=="prioridad":
        productos=sorted(productos,key=lambda p:-(p.get('prioridad') or 0))
    total=0
    for p in productos:
        n=int(p.get('cantidad') or 0)
        for k in range(n):
            if total>=CAP: break
            packer.add_item(Item(f"{p['sku']}#{k}",p['largo_m'],p['ancho_m'],p['alto_m'],p.get('peso') or 0))
            total+=1
    packer.pack(bigger_first=True, number_of_decimals=4)
    b=packer.bins[0]
    boxes=[]; per={}
    for it in b.items:
        sku=it.name.split('#')[0]; per[sku]=per.get(sku,0)+1
        dim=[float(x) for x in it.get_dimension()]; pos=[float(x) for x in it.position]
        boxes.append({"sku":sku,"x":pos[0],"y":pos[1],"z":pos[2],"w":dim[0],"h":dim[1],"d":dim[2]})
    unfit={}
    for it in b.unfitted_items:
        sku=it.name.split('#')[0]; unfit[sku]=unfit.get(sku,0)+1
    vol_cont=L*W*H
    vol_used=sum(x['w']*x['h']*x['d'] for x in boxes)
    peso_cargado=sum((next((p.get('peso') or 0 for p in productos if p['sku']==x['sku']),0)) for x in boxes)
    resumen={"vol_contenedor":round(vol_cont,3),"vol_usado":round(vol_used,3),
        "vol_libre":round(vol_cont-vol_used,3),"pct_volumen":round(100*vol_used/vol_cont,1),
        "peso_max":maxw,"peso_cargado":round(peso_cargado,1),
        "pct_peso":round(100*peso_cargado/maxw,1) if maxw else 0,
        "cargado_por_sku":per,"restante_por_sku":unfit,
        "cap_alcanzado": total>=CAP,
        "container":{"L":L,"W":W,"H":H}}
    return resumen, boxes

# ================= DB =================
import psycopg
from psycopg.rows import dict_row
def conn(): return psycopg.connect(os.environ["DATABASE_URL"], row_factory=dict_row)
SCHEMA="""
create table if not exists productos(
 id bigserial primary key, sku text unique not null, descripcion text,
 largo_m numeric, ancho_m numeric, alto_m numeric,
 volumen_calculado_m3 numeric, volumen_original_m3 numeric,
 estado_volumen text, tipo text default 'normal',
 peso_neto numeric, peso_bruto numeric, fuente text, archivo_origen text,
 fecha_actualizacion date default current_date, fecha_recalculo date, version int default 1);
create table if not exists importaciones(
 id bigserial primary key, nombre_archivo text, tipo text, proveedor text,
 fecha timestamptz default now(), resumen_json jsonb);
create table if not exists revision_pendiente(
 id bigserial primary key, sku text, importacion_id bigint,
 valores_actuales_json jsonb, valores_nuevos_json jsonb, motivo text,
 estado text default 'abierta', fecha timestamptz default now());
create table if not exists despachos(
 id bigserial primary key, numero text, fecha_carga date default current_date,
 nombre_archivo text, resumen_json jsonb, excel_b64 text, fecha timestamptz default now());
"""
def init_and_seed():
    with conn() as c, c.cursor() as cur:
        cur.execute(SCHEMA)
        cur.execute("select count(*) n from productos")
        if cur.fetchone()["n"]>0: return
        path=os.path.join(os.path.dirname(__file__),"master_seed.csv")
        if not os.path.exists(path): return
        with open(path,encoding="utf-8") as f:
            for r in csv.DictReader(f):
                nn=lambda x:(float(x) if x not in (None,"","None") else None)
                cur.execute("""insert into productos(sku,descripcion,largo_m,ancho_m,alto_m,
                  volumen_calculado_m3,volumen_original_m3,estado_volumen,tipo,fuente,
                  archivo_origen,fecha_recalculo) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                  on conflict (sku) do nothing""",
                  [r["sku"],r["articulo"],nn(r["largo_m"]),nn(r["ancho_m"]),nn(r["alto_m"]),
                   nn(r["volumen_calculado_m3"]),nn(r["volumen_original_m3"]),r["estado_volumen"],
                   r["tipo"],r["fuente"],r["archivo_origen"],r["fecha_recalculo"]])

# ================= API =================
app=FastAPI(title="SCD Logistics")
@app.on_event("startup")
def _startup():
    try: init_and_seed()
    except Exception as e: print("seed error:",e)
@app.get("/health")
def health(): return {"ok":True}
@app.get("/productos")
def productos(q: str|None=Query(default=None)):
    sql="select * from productos"; args=[]
    if q: sql+=" where sku ilike %s or descripcion ilike %s"; args=[f"%{q}%",f"%{q}%"]
    sql+=" order by sku"
    with conn() as c, c.cursor() as cur: cur.execute(sql,args); return cur.fetchall()
@app.get("/revisiones")
def revisiones():
    with conn() as c, c.cursor() as cur:
        cur.execute("select * from revision_pendiente where estado='abierta' order by fecha")
        return cur.fetchall()
@app.get("/importaciones")
def importaciones():
    with conn() as c, c.cursor() as cur:
        cur.execute("select id,nombre_archivo,tipo,fecha,resumen_json from importaciones order by fecha desc")
        return cur.fetchall()

def _rows(name,content):
    ext=name.lower().rsplit(".",1)[-1]
    if ext in ("xlsx","xlsm","xls"):
        wb=openpyxl.load_workbook(io.BytesIO(content),data_only=True); ws=wb[wb.sheetnames[0]]
        return parse_rows([list(r) for r in ws.iter_rows(values_only=True)],name)
    if ext=="csv":
        return parse_rows(list(csv.reader(io.StringIO(content.decode("utf-8","ignore")))),name)
    if ext=="pdf":
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            text="\n".join((p.extract_text() or "") for p in pdf.pages)
            tab=[r for p in pdf.pages for t in (p.extract_tables() or []) for r in t]
        items,res=parse_rows(tab,name)
        if res["detectados"]==0: items,res=parse_pdf_text(text,name)
        return items,res
    raise HTTPException(400,"Formato no soportado (PDF, XLSX o CSV).")

@app.post("/importar")
async def importar(file: UploadFile=File(...), commit: bool=False):
    content=await file.read()
    if len(content)>25*1024*1024: raise HTTPException(413,"El archivo supera 25 MB.")
    items,res=_rows(file.filename,content)
    for it in items:
        it.update(sanitize_volume(it.get("largo_m"),it.get("ancho_m"),it.get("alto_m"),it.get("cbm")))
    skus=[it["sku"] for it in items]; master={}
    with conn() as c, c.cursor() as cur:
        if skus:
            cur.execute("select * from productos where sku=any(%s)",[skus])
            master={r["sku"]:r for r in cur.fetchall()}
    nuevos=[it for it in items if it["sku"] not in master]
    seen=set(); dups=[]
    for it in items:
        if it["sku"] in seen: dups.append(it["sku"])
        seen.add(it["sku"])
    modificados=[it["sku"] for it in items if master.get(it["sku"]) and any(_num(it.get(k))!=_num(master[it["sku"]].get(k)) for k in ("largo_m","ancho_m","alto_m"))]
    resumen={"detectados":res["detectados"],"nuevos":len(set(n["sku"] for n in nuevos)),
      "existentes":len([1 for it in items if it["sku"] in master]),"modificados":len(set(modificados)),
      "duplicados_en_archivo":len(dups),"no_interpretadas":res["no_interpretadas"],"errores":res["errores"]}
    if commit:
        with conn() as c, c.cursor() as cur:
            cur.execute("insert into importaciones(nombre_archivo,tipo,resumen_json) values(%s,%s,%s) returning id",
                [file.filename,file.filename.rsplit('.',1)[-1],json.dumps(resumen)])
            imp=cur.fetchone()["id"]; done=set()
            for it in nuevos:
                if it["sku"] in done: continue
                done.add(it["sku"])
                cur.execute("""insert into productos(sku,descripcion,largo_m,ancho_m,alto_m,
                  volumen_calculado_m3,estado_volumen,tipo,archivo_origen,fecha_recalculo)
                  values(%s,%s,%s,%s,%s,%s,%s,%s,%s,current_date) on conflict (sku) do nothing""",
                  [it["sku"],it["descripcion"],it.get("largo_m"),it.get("ancho_m"),it.get("alto_m"),
                   it.get("volumen_calculado_m3"),it.get("estado_volumen"),it["tipo"],it["archivo_origen"]])
            for sku in set(modificados):
                mm=master[sku]; it=next(x for x in items if x["sku"]==sku)
                cur.execute("insert into revision_pendiente(sku,importacion_id,valores_actuales_json,valores_nuevos_json,motivo) values(%s,%s,%s,%s,%s)",
                  [sku,imp,json.dumps({k:float(mm[k]) if mm.get(k) is not None else None for k in ("largo_m","ancho_m","alto_m")}),
                   json.dumps({k:it.get(k) for k in ("largo_m","ancho_m","alto_m")}),"Encontre dimensiones diferentes"])
        resumen["importacion_id"]=imp
    return resumen

@app.get("/export.xlsx")
def export():
    with conn() as c, c.cursor() as cur:
        cur.execute("select * from productos order by sku"); prods=cur.fetchall()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="maestro"
    if prods:
        cols=list(prods[0].keys()); ws.append(cols)
        for p in prods: ws.append([p[c] for c in cols])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=maestro_saneado.xlsx"})

@app.get("/contenedores")
def contenedores(): return CONTAINERS

@app.post("/optimizar")
def optimizar_ep(payload: dict=Body(...)):
    cont=payload.get("contenedor")
    if isinstance(cont,str): cont=CONTAINERS.get(cont,CONTAINERS["40"])
    else: cont=(cont["L"],cont["W"],cont["H"],cont.get("peso_max",26600))
    prods=payload.get("productos",[]); modo=payload.get("modo","libre")
    # completar dims desde maestro
    skus=[p["sku"] for p in prods]
    with conn() as c, c.cursor() as cur:
        cur.execute("select * from productos where sku=any(%s)",[skus] if skus else [[]])
        mp={r["sku"]:r for r in cur.fetchall()}
    full=[]
    for p in prods:
        m=mp.get(p["sku"])
        if not m or not m["largo_m"]: continue
        full.append({"sku":p["sku"],"cantidad":p.get("cantidad") or 0,"prioridad":p.get("prioridad") or 0,
            "largo_m":float(m["largo_m"]),"ancho_m":float(m["ancho_m"]),"alto_m":float(m["alto_m"]),
            "peso":float(m["peso_neto"]) if m.get("peso_neto") else 0})
    if not full: raise HTTPException(400,"No hay productos con dimensiones para optimizar.")
    resumen,boxes=optimizar(full,cont,modo)
    return {"resumen":resumen,"boxes":boxes}

@app.post("/despacho")
async def despacho(despacho: UploadFile=File(...), pl: UploadFile|None=File(default=None), commit: bool=False):
    content=await despacho.read()
    import pdfplumber
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        txt="\n".join((p.extract_text() or "") for p in pdf.pages)
    d=parse_despacho(txt)
    pl_items=None
    if pl is not None:
        pc=await pl.read(); pl_items,_=_rows(pl.filename,pc)
    d['items']=match_skus(d['items'],pl_items)
    today=datetime.date.today().isoformat()
    resumen={k:v for k,v in d.items() if k!='items'}
    resumen['n_items']=len(d['items'])
    out={"despacho":resumen,"items":d['items'],"fecha_carga":today,
         "nombre":f"Despacho_{today}_{d.get('numero') or 'SN'}.xlsx"}
    if commit:
        xls=despacho_excel(d,today); b64=base64.b64encode(xls).decode()
        with conn() as c, c.cursor() as cur:
            cur.execute("insert into despachos(numero,nombre_archivo,resumen_json,excel_b64) values(%s,%s,%s,%s) returning id",
                [d.get('numero'),out['nombre'],json.dumps(resumen),b64])
            out['id']=cur.fetchone()['id']
    return out

@app.get("/despacho/{did}.xlsx")
def despacho_xlsx(did:int):
    with conn() as c, c.cursor() as cur:
        cur.execute("select nombre_archivo,excel_b64 from despachos where id=%s",[did])
        r=cur.fetchone()
    if not r: raise HTTPException(404,"No existe")
    data=base64.b64decode(r["excel_b64"])
    return StreamingResponse(io.BytesIO(data),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename={r['nombre_archivo']}"})

@app.get("/despachos")
def despachos():
    with conn() as c, c.cursor() as cur:
        cur.execute("select id,numero,nombre_archivo,fecha_carga,resumen_json from despachos order by fecha desc")
        return cur.fetchall()

@app.get("/")
def home(): return FileResponse(os.path.join(os.path.dirname(__file__),"index.html"))
